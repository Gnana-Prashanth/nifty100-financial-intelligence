import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -----------------------------
# Load financial_ratios
# -----------------------------
conn = sqlite3.connect("nifty100.db")

df = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

conn.close()

# -----------------------------
# KPI Columns (Actual Project Columns)
# -----------------------------
kpi_columns = [
    "company_id",
    "year_x",
    "composite_quality_score",
    "sector_relative_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "sales"
]

# Keep only columns that actually exist
kpi_columns = [c for c in kpi_columns if c in df.columns]

# -----------------------------
# Screener Files
# -----------------------------
preset_files = {
    "Quality": "output/quality_screener.csv",
    "Growth": "output/growth_screener.csv",
    "Value": "output/value_screener.csv",
    "Dividend": "output/dividend_screener.csv",
    "Turnaround": "output/turnaround_screener.csv",
    "Compounder": "output/compounder_screener.csv"
}

output_file = "output/screener_output.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    for sheet_name, csv_file in preset_files.items():

        if not os.path.exists(csv_file):
            print(f"{csv_file} not found. Skipping...")
            continue

        preset = pd.read_csv(csv_file)

        # Use the screener CSV directly
        export = preset.copy()

        # Keep only the columns that actually exist
        available_columns = [
            col for col in kpi_columns
            if col in export.columns
        ]

        # Sort if the score exists
        if "composite_quality_score" in export.columns:
            export = export.sort_values(
                "composite_quality_score",
                ascending=False
            )

        export[available_columns].to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )
        
# -----------------------------
# Conditional Formatting
# -----------------------------
wb = load_workbook(output_file)

green = PatternFill(
    fill_type="solid",
    start_color="90EE90"
)

red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)

for ws in wb.worksheets:

    headers = {}

    for cell in ws[1]:
        headers[cell.value] = cell.column

    for row in range(2, ws.max_row + 1):

        # ROE
        if "return_on_equity_pct" in headers:

            col = headers["return_on_equity_pct"]

            value = ws.cell(row=row, column=col).value

            if value is not None:
                ws.cell(row=row, column=col).fill = green if value >= 15 else red

        # Debt to Equity
        if "debt_to_equity" in headers:

            col = headers["debt_to_equity"]

            value = ws.cell(row=row, column=col).value

            if value is not None:
                ws.cell(row=row, column=col).fill = green if value <= 1 else red

        # Interest Coverage
        if "interest_coverage" in headers:

            col = headers["interest_coverage"]

            value = ws.cell(row=row, column=col).value

            if value is not None:
                ws.cell(row=row, column=col).fill = green if value >= 3 else red

wb.save(output_file)

print("Excel report generated successfully!")