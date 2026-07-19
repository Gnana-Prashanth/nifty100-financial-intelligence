import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ----------------------------------------------------
# Database Connection
# ----------------------------------------------------

conn = sqlite3.connect("nifty100.db")

companies = pd.read_sql(
    """
    SELECT
        id,
        company_name
    FROM companies
    """,
    conn
)

peer_groups = pd.read_sql(
    """
    SELECT *
    FROM peer_groups
    """,
    conn
)

financial_ratios = pd.read_sql(
    """
    SELECT *
    FROM financial_ratios
    """,
    conn
)

peer_percentiles = pd.read_sql(
    """
    SELECT *
    FROM peer_percentiles
    """,
    conn
)

# ----------------------------------------------------
# Merge Company Names
# ----------------------------------------------------

peer_groups = peer_groups.merge(
    companies,
    left_on="company_id",
    right_on="id",
    how="left"
)

peer_groups.drop(columns=["id_y"], inplace=True)
peer_groups.rename(columns={"id_x": "id"}, inplace=True)

# ----------------------------------------------------
# Metric Mapping
# ----------------------------------------------------

metric_mapping = {
    "ROE": "return_on_equity_pct",
    "ROCE": "return_on_capital_employed_pct",
    "Net Profit Margin": "net_profit_margin_pct",
    "Debt to Equity": "debt_to_equity",
    "Free Cash Flow": "free_cash_flow_cr",
    "PAT CAGR 5yr": "pat_cagr_5yr",
    "Revenue CAGR 5yr": "revenue_cagr_5yr",
    "EPS CAGR 5yr": "eps_cagr_5yr",
    "Interest Coverage": "interest_coverage",
    "Asset Turnover": "asset_turnover"
}

# ----------------------------------------------------
# Workbook
# ----------------------------------------------------

wb = Workbook()

wb.remove(wb.active)

# ----------------------------------------------------
# Styles
# ----------------------------------------------------

header_fill = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

green_fill = PatternFill(
    fill_type="solid",
    start_color="92D050",
    end_color="92D050"
)

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="F4CCCC",
    end_color="F4CCCC"
)

benchmark_fill = PatternFill(
    fill_type="solid",
    start_color="FFC000",
    end_color="FFC000"
)

summary_fill = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

peer_group_names = sorted(
    peer_groups["peer_group_name"].unique()
)

print("Peer Groups:")

for group in peer_group_names:
    print("-", group)

# ----------------------------------------------------
# Create Worksheets & Populate Data
# ----------------------------------------------------

for peer_group in peer_group_names:

    print(f"Creating sheet : {peer_group}")

    ws = wb.create_sheet(title=peer_group[:31])

    headers = ["Company ID", "Company Name"]

    for metric in metric_mapping.keys():
        headers.append(metric)
        headers.append(f"{metric} Percentile")

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    peer_companies = peer_groups[
        peer_groups["peer_group_name"] == peer_group
    ]

    excel_row = 2

    for _, company in peer_companies.iterrows():

        company_id = company["company_id"]

        ws.cell(row=excel_row, column=1).value = company_id
        ws.cell(row=excel_row, column=2).value = company["company_name"]

        # -----------------------------
        # Latest Financial Ratios
        # -----------------------------

        fr = financial_ratios[
            financial_ratios["company_id"] == company_id
        ].copy()

        if fr.empty:
            excel_row += 1
            continue

        non_ttm = fr[
            fr["year"] != "TTM"
        ]

        if not non_ttm.empty:
            latest_year = sorted(non_ttm["year"])[-1]
            fr = non_ttm[
                non_ttm["year"] == latest_year
            ]
        else:
            fr = fr[
                fr["year"] == "TTM"
            ]

        fr = fr.iloc[0]

        # -----------------------------
        # Latest Percentiles
        # -----------------------------

        pp = peer_percentiles[
            (peer_percentiles["company_id"] == company_id) &
            (peer_percentiles["peer_group"] == peer_group)
        ].copy()

        if not pp.empty:

            non_ttm_pp = pp[
                pp["year"] != "TTM"
            ]

            if not non_ttm_pp.empty:
                latest_pp_year = sorted(non_ttm_pp["year"])[-1]
                pp = non_ttm_pp[
                    non_ttm_pp["year"] == latest_pp_year
                ]
            else:
                pp = pp[
                    pp["year"] == "TTM"
                ]

        current_col = 3

        # -----------------------------
        # Write Metrics
        # -----------------------------

        for metric_name, column_name in metric_mapping.items():

            value = fr.get(column_name, None)

            percentile = None

            if not pp.empty:

                temp = pp[
                    pp["metric"] == metric_name
                ]

                if not temp.empty:
                    percentile = temp.iloc[0]["percentile_rank"]

            ws.cell(
                row=excel_row,
                column=current_col
            ).value = value

            ws.cell(
                row=excel_row,
                column=current_col + 1
            ).value = percentile

            current_col += 2

        # -----------------------------
        # Highlight Benchmark Company
        # -----------------------------

        if company["is_benchmark"] == 1:

            for c in range(1, len(headers) + 1):
                ws.cell(
                    row=excel_row,
                    column=c
                ).fill = benchmark_fill

        excel_row += 1

# ----------------------------------------------------
# Apply Formatting
# ----------------------------------------------------

for ws in wb.worksheets:

    last_row = ws.max_row
    last_col = ws.max_column

    # ----------------------------------------
    # Colour Percentile Columns
    # ----------------------------------------

    for row in range(2, last_row + 1):

        for col in range(4, last_col + 1, 2):

            cell = ws.cell(row=row, column=col)

            if cell.value is None:
                continue

            try:
                value = float(cell.value)
            except:
                continue

            if value >= 75:
                cell.fill = green_fill

            elif value >= 25:
                cell.fill = yellow_fill

            else:
                cell.fill = red_fill

    # ----------------------------------------
    # Median Summary Row
    # ----------------------------------------

    summary_row = last_row + 1

    ws.cell(summary_row, 1).value = "Median"
    ws.cell(summary_row, 1).font = Font(bold=True)

    for c in range(1, last_col + 1):
        ws.cell(summary_row, c).fill = summary_fill

    col = 3

    while col <= last_col:

        values = []

        for r in range(2, last_row + 1):

            value = ws.cell(r, col).value

            if isinstance(value, (int, float)):
                values.append(value)

        if values:

            ws.cell(
                row=summary_row,
                column=col
            ).value = round(pd.Series(values).median(), 2)

        col += 2

    # ----------------------------------------
    # Auto Width
    # ----------------------------------------

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

# ----------------------------------------------------
# Save Workbook
# ----------------------------------------------------

import os

os.makedirs("output", exist_ok=True)

output_file = os.path.join(
    "output",
    "peer_comparison.xlsx"
)

wb.save(output_file)

conn.close()

print("=" * 60)
print("Peer Comparison Excel Report Generated Successfully")
print(f"Saved to: {output_file}")
print("=" * 60)